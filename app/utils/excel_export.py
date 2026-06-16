from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill


def export_quiz_results(students_data, quiz_title):
    wb = Workbook()
    ws = wb.active
    ws.title = "Kết quả bài tập"

    headers = ["STT", "Học sinh", "Email", "Điểm", "Câu đúng", "Câu sai", "Thời gian làm (phút)", "Ngày nộp"]
    header_fill = PatternFill(start_color="22C55E", end_color="22C55E", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for idx, row in enumerate(students_data, 1):
        ws.cell(row=idx + 1, column=1, value=idx)
        ws.cell(row=idx + 1, column=2, value=row.get("student_name", ""))
        ws.cell(row=idx + 1, column=3, value=row.get("email", ""))
        ws.cell(row=idx + 1, column=4, value=row.get("score", 0))
        ws.cell(row=idx + 1, column=5, value=row.get("correct", 0))
        ws.cell(row=idx + 1, column=6, value=row.get("wrong", 0))
        ws.cell(row=idx + 1, column=7, value=row.get("time_taken", 0))
        ws.cell(row=idx + 1, column=8, value=row.get("submitted_at", ""))

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_students_list(students, class_name=""):
    wb = Workbook()
    ws = wb.active
    ws.title = "Danh sách học sinh"

    headers = ["STT", "Họ tên", "Email", "Lớp", "Trạng thái", "Ngày tạo"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    for idx, s in enumerate(students, 1):
        ws.cell(row=idx + 1, column=1, value=idx)
        ws.cell(row=idx + 1, column=2, value=s.get("full_name", ""))
        ws.cell(row=idx + 1, column=3, value=s.get("email", ""))
        ws.cell(row=idx + 1, column=4, value=class_name or s.get("class_name", ""))
        ws.cell(row=idx + 1, column=5, value=s.get("status", ""))
        ws.cell(row=idx + 1, column=6, value=s.get("created_at", ""))

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
